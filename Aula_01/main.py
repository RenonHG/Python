"""
A proposta das aulas será desenvolver um campo minado

Básico de python
"""
print('''\033[35m
      
██████  ██    ██ ████████ ██   ██  ██████  ███    ██ 
██   ██  ██  ██     ██    ██   ██ ██    ██ ████   ██ 
██████    ████      ██    ███████ ██    ██ ██ ██  ██ 
██         ██       ██    ██   ██ ██    ██ ██  ██ ██ 
██         ██       ██    ██   ██  ██████  ██   ████ 
                                                     
\033[0;0m''')



# print('Bem vindo ao sistema malvadão: ')
# # numero1 = input('Digite um número: ')
# nome = 'Renon'
# idade = 18
# endereco = 'rua dos bobos'
# numero = 0
# telefone = '19 99999-9999'

# 'Dados cadastrados com sucesso'

# # print("\nCaro Sr. " + nome + " com a idade de " + idade + " anos, " +
# #  "residente na " + endereco + ", " + numero + ". Telefone: " + telefone)

# print("\nCaro Sr. {} com a idade de  {} anos, residente na {}, Nº {}. Telefone: {}"
#       .format(nome, idade, endereco, numero, telefone))
# #---------------------------------------------

# # Estrutura do IF, ELIF e ELSE
# if idade >= 18:
#     print('Você é maior de idade')
# elif idade < 10:
#     print('Você tem menos de 10 anos')
# else:
#     print('Você NÃO é maior de idade nem menor que 10 anos!')
# #---------------------------------------------

# # Estrutura do in
# vogal = ['a', 'e', 'i', 'o', 'u']

# if 'i' in vogal:
#     print('é vogal!')
# else:
#     print('Não é vogal!')
# #---------------------------------------------

# # Estrutura do For
# for i in range(1, 6):
#     print(str(i) + '- Calma calabreso!')

# # for e in juntos para listar cada letra na string
# for letra in '- Calma calabreso!':
#     print(letra)

# # for e in juntos para listar cada letra na vogal
# for letra in vogal:
#     print(letra)

#----------------------------------------------

# '''
# 01- Atividade
#     O sistema deve solicitar o número ao usuario
#     O sistema deve percorrer do 0 até o número escolhido 🆗
#     Se o numero for menor ou igual do 0 peça que o usuário digite novamente até receber um número valido 🆗
#     No laço se o número for PAR deve ficar em AZUL e se for IMPAR deve ficar em VERMELHO

# '''
# # numero = input('Digite um número maior que 0: ')
# numero = '0'

# while numero < '1':
#     numero = input('Digite um número maior que 0: ')


# for i in range(0, int(numero)+1):
#     if i % 2 == 0:
#         print('\033[34m' + str(i))
#     else:
#         print('\033[31m' + str(i))
# ---------------------------------------------


'''
02 - Atividade
    -Solicitar ao usuário um digito 🆗
    - Deve apresentar a quantidade em * na tela 🆗
    - solicitar 2º número 🆗
    - deve ser a quantidade de linhas apresentadas na tela 🆗
    
    - Crie o menu de acesso do jogo com as seguites opções: Jogar, Configurar e Sair 🆗
    - Hoje a gente vai criar o configurar, dê as opções ao usuário para ele configurar o jogo dele
    - Nessas configurações coloque uma terceira config "dificuldade", o usuário vai poder escolher  entre 1 e 3

'''
from openpyxl import Workbook, load_workbook
import random
 

def configs():
    try:
        wb = load_workbook(filename='campo.xlsx')
        config = wb['config']
        
    except:
        #criando tabela 
        wb = Workbook()
        config = wb.create_sheet('config')
        config.cell(column=1, row=1, value="linha")
        config.cell(column=2, row=1, value=5)
        config.cell(column=1, row=2, value="coluna")
        config.cell(column=2, row=2, value=5)
        config.cell(column=1, row=3, value="dificuldade")
        config.cell(column=2, row=3, value=2)


    linhas = config.cell(column=2, row=1).value
    colunas = config.cell(column=2, row=2).value
    dificuldade = config.cell(column=2, row=3).value



    wb.save('campo.xlsx')

    config = {
        'linhas' : int(linhas),
        'colunas' : int(colunas),
        'dificuldade' : int(dificuldade)
    }



    return config

def criarTabuleiro():
    config = configs()
    qtdLinha = config['linhas']
    qtdColuna = config['colunas']
    bombas = calculoBombas()
    tabuleiro = []
    cont = 1

    #0 → Casa vazia |     -1 → bombas 

    for i in range(0, int(qtdLinha)):
        linha = []
        for j in range(0, int(qtdColuna)):
            print('{:5}' . format('[  ]'), end=' ')
            if cont in bombas:
                linha.append(-1)
                # linha.append('💣')
            else:
                linha.append(0)

            cont += 1
        print()
        # print('- - - ' * int(qtdColuna))
        tabuleiro.append(linha)

    gravarTabuleiro(tabuleiro)

def gravarTabuleiro(tabuleiro):
    try:
        wb = load_workbook('campo.xlsx')
    except:
        Workbook()

    try:
        abaJogo = wb['jogo']
    except:
        abaJogo = wb.create_sheet('jogo')

    abaJogo.delete_rows(1, abaJogo.max_row)
    abaJogo.delete_cols(1, abaJogo.max_column)

    for linha in range(1, len(tabuleiro) + 1):
        for coluna in range(1, len(tabuleiro[0]) + 1):
            abaJogo.cell(row=linha, column=coluna, value=tabuleiro[linha -1][coluna -1])

    wb.save('campo.xlsx')
    
def configCampo():
    # novaQtdColuna = 3
    # novaQtdLinha = 3
    # novaDificuldade = 1

    #LINHAS 
    novaQtdLinha = input('Nº de Linhas (mínimo 3): ')
    while int(novaQtdLinha) < 3:
        print("Valor inválido, insira novamente um valor maior ou igual a 3!")
        novaQtdLinha = input('Quntas Linhas : ')
    
    #COLUNAS 
    novaQtdColuna = input('Nº de Colunas (mínimo 3): ')
    while int(novaQtdColuna) < 3:
        print("Valor inválido, insira novamente um valor maior ou igual a 3!")
        novaQtdColuna = input('Quntas Colunas : ')

    #DIFICULDADE 
    novaDificuldade = input('1- Fácil \n2- Médio \n3- Difícil \n\nEscolha a dificuldade: ')
    while int(novaDificuldade) < 1 or int(novaDificuldade) > 3:
        print("Valor inválido, escolha novamente!")
        novaDificuldade = input('1- Fácil \n2- Médio \n3- Difícil \n\nEscolha a dificuldade: ')

    try:
        wb = load_workbook(filename='campo.xlsx')
        config = wb['config']
    except:
        wb = Workbook()
        config = wb.create_sheet('config')

    config.cell(column=1, row=1, value="linha")
    config.cell(column=2, row=1, value=novaQtdLinha)
    config.cell(column=1, row=2, value="coluna")
    config.cell(column=2, row=2, value=novaQtdColuna)
    config.cell(column=1, row=3, value="dificuldade")
    config.cell(column=2, row=3, value=novaDificuldade)

    wb.save('campo.xlsx')

def calculoBombas():
    '''
    Descubra o valor total de 'casas' no tabuleiro, e a partir desse total calcule a quantidade de bombas que vai ter no tabuleiro;
    fácil → 15%, médio → 30% e díficil → 50%. 
    '''
    config = configs()
    totalCasas = config['linhas'] * config['colunas']

    
    if config['dificuldade'] == 1:
        quantBomba = totalCasas * 0.15

    if config['dificuldade'] == 2:
        quantBomba = totalCasas * 0.30

    if config['dificuldade'] == 3:
        quantBomba = totalCasas * 0.50

    # print('Total de bombas: {}' .format(int(quantBomba)))

    bombas = []
    while True:
        posicao = random.randint(1, totalCasas)

        if posicao not in bombas:
            bombas.append(posicao)

        if len(bombas) == int(quantBomba):
            break

    # print(sorted(bombas))
        # print("Bomba na casa: {}"  .format(int(random.randint(1, 10))))

    #Aleatoriedade das bombinhas do mal:
    return bombas

def jogar():
    criarTabuleiro()
    wb          = load_workbook(filename='campo.xlsx')
    jogo        = wb['jogo']
    maxLinha    = jogo.max_row
    maxColuna   = jogo.max_column
    gameOver    = False
    jgPossivel = (maxLinha * maxColuna) - len(calculoBombas())
    historico = 0

    while True:
        # linhaJogada = 3
        # colunaJogada = 2
        linhaJogada = int(input('Linha: '))
        while linhaJogada > maxLinha:
            print("Esse valor excede o número de linhas, jogue novamente um valor menor: ")
            linhaJogada = int(input('Linha: '))
                
        colunaJogada = int(input('Coluna: '))
        while colunaJogada > maxColuna:
            print("Esse valor excede o número de colunas, jogue novamente um valor menor: ")
            colunaJogada = int(input('Coluna: '))


        # [ ], 🆗, 💣

        jogada = int(jogo.cell(row=linhaJogada, column=colunaJogada).value)

        if jogada == 0:
            jogo.cell(row=linhaJogada, column=colunaJogada, value=1)
            historico += 1
        elif jogada == -1:
            # jogo.cell(row=linhaJogada, column=colunaJogada, value=-2)
            gameOver = True
        elif jogada == 1:
            print('Você ja tentou esse, tente novamente!')

        wb.save('campo.xlsx')
        for linha in range(1, maxLinha + 1):
            for coluna in range(1, maxColuna + 1):
                casa = jogo.cell(row=linha, column=coluna).value
                if int(casa) == 0 or (int(casa) == -1 and gameOver == False):
                    print('{:5}' . format('[🏠]'), end='')
                elif int(casa) == 1:
                    print('{:5}' . format('[🆗]'), end='')
                elif int(casa) == -1 and gameOver == True:
                    print('{:5}' . format('[💥]'), end='')

            print()
        if gameOver == True:
            print("Allahu Akbar 💥💥💥💥❗❗❗")
            break
        if historico == jgPossivel:
            print("Parabéns, você ganhou❗ 🎉")
            break
    if historico == jgPossivel or gameOver == True:
        op = input('Deseja jogar novamente? (s/n):')
        if op == 's':
            jogar()
        


#============================ MENU ========================================

while True:
    #MENU DE SELEÇÃO
    print('1 - JOGAR \n2 - CONFIGURAR \n3 - SAIR')
    
    opcao = input('Escolha uma opção: ')

    if opcao == '1':
        jogar()

    elif opcao == '2':
        configCampo()
        
    elif opcao == '3':
        print('Sair')
        break
                
    else:
        print('Selecione uma opção válida!')


